"""Aggregate reporting: late/early-leave rankings and absence reports.

Every report here is scoped the same way as ``/api/logs``: a campus director
only ever sees their own campus; hq sees everything (optionally filtered to
one campus). "Late" / "early leave" are computed against each staff member's
*campus* shift hours (``Campus.shift_start`` / ``shift_end``, settable only by
hq) plus a caller-supplied ``threshold_minutes`` grace window.
"""
import calendar
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..config import settings
from ..database import get_db
from ..deps import get_current_manager
from ..models import (
    AttendanceLog,
    AttendanceType,
    Campus,
    Holiday,
    LeaveRecord,
    LeaveStatus,
    LocationViolation,
    User,
    UserRole,
    UserStatus,
    ensure_aware,
)
from ..schemas import (
    AbsenceDayEntry,
    AbsenceReasonStat,
    AbsenceSummaryResponse,
    AbsenceTotalEntry,
    DailyTrendEntry,
    DailyTrendResponse,
    EarlyLeaveEntry,
    EarlyLeaveRankingEntry,
    ForgotCheckoutEntry,
    ForgotCheckoutResponse,
    LateArrivalEntry,
    LateRankingEntry,
    LocationAlertEntry,
    LocationAlertsResponse,
    MonthlyHoursEntry,
    MonthlyHoursResponse,
    RiskReportResponse,
    RiskStaffEntry,
    TodayAbsenteeEntry,
    TodayAbsenteesResponse,
    UnresolvedReminderResponse,
)
from ..scoping import scope_campus_id
from ..services import effective_working_days

router = APIRouter(prefix="/api/reports", tags=["reports"])

MAX_REPORT_DAYS = 400


def _validate_range(start_date: date, end_date: date) -> None:
    if end_date < start_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="end_date, start_date'den önce olamaz.")
    if (end_date - start_date).days > MAX_REPORT_DAYS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tarih aralığı en fazla {MAX_REPORT_DAYS} gün olabilir.",
        )


def _day_range(start_date: date, end_date: date, exclude_weekends: bool):
    d = start_date
    while d <= end_date:
        if not (exclude_weekends and d.weekday() >= 5):
            yield d
        d += timedelta(days=1)


def _all_days(start_date: date, end_date: date) -> list[date]:
    out: list[date] = []
    d = start_date
    while d <= end_date:
        out.append(d)
        d += timedelta(days=1)
    return out


async def _load_holidays(
    db: AsyncSession, start_date: date, end_date: date
) -> tuple[set[date], dict[int, set[date]]]:
    """Holidays overlapping the range, split into national (all-campus) dates
    and a per-campus map of campus-scoped closure dates."""
    rows = await db.execute(
        select(Holiday).where(Holiday.date >= start_date, Holiday.date <= end_date)
    )
    national: set[date] = set()
    by_campus: dict[int, set[date]] = defaultdict(set)
    for h in rows.scalars().all():
        if h.campus_id is None:
            national.add(h.date)
        else:
            by_campus[h.campus_id].add(h.date)
    return national, by_campus


def _tracking_start(staff: User) -> date:
    """The first day this staff member is counted for attendance: the later of
    the global go-live date and their own registration (account-creation) date.

    Days before this are never expected and never absences — an existing staff
    member isn't penalised for the period before go-live, and someone who
    registers later isn't penalised for the days before they joined.
    """
    go_live = settings.go_live_date
    created = staff.created_at
    if created is None:
        return go_live
    tz = ZoneInfo(settings.attendance_timezone)
    created_local = ensure_aware(created).astimezone(tz).date()
    return max(created_local, go_live)


def _expected_days_for_staff(
    staff: User,
    all_days: list[date],
    exclude_weekends: bool,
    national_holidays: set[date],
    holidays_by_campus: dict[int, set[date]],
) -> list[date]:
    """The days in the range a staff member is actually expected to work:
    their per-person working weekdays, minus any applicable holiday, and never
    before their attendance tracking start (go-live / registration date)."""
    working = effective_working_days(staff.working_days, exclude_weekends)
    closed = national_holidays | holidays_by_campus.get(staff.campus_id, set())
    start = _tracking_start(staff)
    return [
        d
        for d in all_days
        if d >= start and d.isoweekday() in working and d not in closed
    ]


async def _scoped_active_staff(
    db: AsyncSession, manager: User, campus_id: int | None, user_id: int | None
) -> list[User]:
    scope = scope_campus_id(manager, campus_id)
    stmt = select(User).where(User.role == UserRole.staff, User.status == UserStatus.active)
    if scope is not None:
        stmt = stmt.where(User.campus_id == scope)
    if user_id is not None:
        stmt = stmt.where(User.id == user_id)
    return list((await db.execute(stmt)).scalars().all())


async def _logs_for_staff(
    db: AsyncSession, staff_ids: list[int], start_date: date, end_date: date
) -> list[AttendanceLog]:
    if not staff_ids:
        return []
    tz = ZoneInfo(settings.attendance_timezone)
    start_utc = datetime.combine(start_date, time.min, tzinfo=tz).astimezone(timezone.utc)
    end_utc = datetime.combine(end_date, time.max, tzinfo=tz).astimezone(timezone.utc)
    stmt = select(AttendanceLog).where(
        AttendanceLog.user_id.in_(staff_ids),
        AttendanceLog.scan_time >= start_utc,
        AttendanceLog.scan_time <= end_utc,
    )
    return list((await db.execute(stmt)).scalars().all())


@dataclass
class _DayLogs:
    first_in: datetime | None = None
    last_out: datetime | None = None
    any_log: bool = False


def _group_by_staff_day(logs: list[AttendanceLog], tz: ZoneInfo) -> dict[tuple[int, date], _DayLogs]:
    grouped: dict[tuple[int, date], _DayLogs] = defaultdict(_DayLogs)
    for log in logs:
        local_dt = log.scan_time.astimezone(tz)
        key = (log.user_id, local_dt.date())
        bucket = grouped[key]
        bucket.any_log = True
        if log.type == AttendanceType.IN and (bucket.first_in is None or local_dt < bucket.first_in):
            bucket.first_in = local_dt
        if log.type == AttendanceType.OUT and (bucket.last_out is None or local_dt > bucket.last_out):
            bucket.last_out = local_dt
    return grouped


async def _campus_names_map(db: AsyncSession) -> dict[int, str]:
    rows = await db.execute(select(Campus.id, Campus.name))
    return {cid: name for cid, name in rows.all()}


async def _campus_shift_map(db: AsyncSession) -> dict[int, Campus]:
    rows = await db.execute(select(Campus))
    return {c.id: c for c in rows.scalars().all()}


@router.get("/today-absentees", response_model=TodayAbsenteesResponse)
async def today_absentees(
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    exclude_weekends: bool = True,
):
    """Who is expected today but not here yet: active staff whose working
    schedule includes today (and who are past their tracking start), with no
    scan today and no active leave covering today. This is the daily "who hasn't
    come in" list. During the day it naturally shrinks as people scan in."""
    tz = ZoneInfo(settings.attendance_timezone)
    today = datetime.now(timezone.utc).astimezone(tz).date()
    staff = await _scoped_active_staff(db, manager, campus_id, None)
    if not staff:
        return TodayAbsenteesResponse(date=today, count=0, entries=[])

    staff_ids = [s.id for s in staff]
    logs = await _logs_for_staff(db, staff_ids, today, today)
    present = {log.user_id for log in logs}  # any scan today = here

    leave_rows = await db.execute(
        select(LeaveRecord.user_id).where(
            LeaveRecord.user_id.in_(staff_ids),
            LeaveRecord.status == LeaveStatus.active,
            LeaveRecord.start_date <= today,
            LeaveRecord.end_date >= today,
        )
    )
    on_leave = {uid for (uid,) in leave_rows.all()}

    national_holidays, holidays_by_campus = await _load_holidays(db, today, today)
    campus_names = await _campus_names_map(db)

    entries: list[TodayAbsenteeEntry] = []
    for s in staff:
        # Empty when today isn't a working day for them, is a holiday, or is
        # before their go-live/registration tracking start.
        if not _expected_days_for_staff(
            s, [today], exclude_weekends, national_holidays, holidays_by_campus
        ):
            continue
        if s.id in present or s.id in on_leave:
            continue
        entries.append(
            TodayAbsenteeEntry(
                user_id=s.id,
                full_name=s.full_name,
                job_title=s.job_title,
                branch=s.branch,
                campus_name=campus_names.get(s.campus_id) if s.campus_id else None,
                phone=s.phone,
            )
        )

    entries.sort(key=lambda e: ((e.campus_name or ""), e.full_name))
    return TodayAbsenteesResponse(date=today, count=len(entries), entries=entries)


@router.get("/late", response_model=list[LateRankingEntry])
async def late_ranking(
    start_date: date,
    end_date: date,
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    threshold_minutes: int = Query(0, ge=0, le=240, description="Grace period before counting as late"),
    exclude_weekends: bool = True,
    limit: int = Query(50, ge=1, le=700),
):
    """Ranking of staff by how often their first IN of the day is later than
    the campus shift start + grace period."""
    _validate_range(start_date, end_date)
    tz = ZoneInfo(settings.attendance_timezone)
    staff = await _scoped_active_staff(db, manager, campus_id, user_id)
    if not staff:
        return []

    logs = await _logs_for_staff(db, [s.id for s in staff], start_date, end_date)
    grouped = _group_by_staff_day(logs, tz)
    campuses = await _campus_shift_map(db)
    national_holidays, holidays_by_campus = await _load_holidays(db, start_date, end_date)
    all_days = _all_days(start_date, end_date)

    results: list[LateRankingEntry] = []
    for s in staff:
        campus = campuses.get(s.campus_id) if s.campus_id else None
        if campus is None or campus.shift_start is None:
            continue  # no shift configured for this campus — cannot judge lateness
        late_minutes: list[float] = []
        for d in _expected_days_for_staff(
            s, all_days, exclude_weekends, national_holidays, holidays_by_campus
        ):
            bucket = grouped.get((s.id, d))
            if bucket is None or bucket.first_in is None:
                continue
            shift_start_local = datetime.combine(d, campus.shift_start, tzinfo=tz)
            minutes_late = (bucket.first_in - shift_start_local).total_seconds() / 60
            if minutes_late > threshold_minutes:
                late_minutes.append(minutes_late)
        if late_minutes:
            results.append(
                LateRankingEntry(
                    user_id=s.id,
                    full_name=s.full_name,
                    job_title=s.job_title,
                    branch=s.branch,
                    campus_name=campus.name,
                    late_days=len(late_minutes),
                    average_late_minutes=round(sum(late_minutes) / len(late_minutes), 1),
                )
            )

    results.sort(key=lambda r: (-r.late_days, -r.average_late_minutes))
    return results[:limit]


@router.get("/early-leave", response_model=list[EarlyLeaveRankingEntry])
async def early_leave_ranking(
    start_date: date,
    end_date: date,
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    threshold_minutes: int = Query(0, ge=0, le=240, description="Grace period before counting as early"),
    exclude_weekends: bool = True,
    limit: int = Query(50, ge=1, le=700),
):
    """Ranking of staff by how often their last OUT of the day is earlier than
    the campus shift end - grace period."""
    _validate_range(start_date, end_date)
    tz = ZoneInfo(settings.attendance_timezone)
    staff = await _scoped_active_staff(db, manager, campus_id, user_id)
    if not staff:
        return []

    logs = await _logs_for_staff(db, [s.id for s in staff], start_date, end_date)
    grouped = _group_by_staff_day(logs, tz)
    campuses = await _campus_shift_map(db)
    national_holidays, holidays_by_campus = await _load_holidays(db, start_date, end_date)
    all_days = _all_days(start_date, end_date)

    results: list[EarlyLeaveRankingEntry] = []
    for s in staff:
        campus = campuses.get(s.campus_id) if s.campus_id else None
        if campus is None or campus.shift_end is None:
            continue
        early_minutes: list[float] = []
        for d in _expected_days_for_staff(
            s, all_days, exclude_weekends, national_holidays, holidays_by_campus
        ):
            bucket = grouped.get((s.id, d))
            if bucket is None or bucket.last_out is None:
                continue
            shift_end_local = datetime.combine(d, campus.shift_end, tzinfo=tz)
            minutes_early = (shift_end_local - bucket.last_out).total_seconds() / 60
            if minutes_early > threshold_minutes:
                early_minutes.append(minutes_early)
        if early_minutes:
            results.append(
                EarlyLeaveRankingEntry(
                    user_id=s.id,
                    full_name=s.full_name,
                    job_title=s.job_title,
                    branch=s.branch,
                    campus_name=campus.name,
                    early_leave_days=len(early_minutes),
                    average_early_minutes=round(sum(early_minutes) / len(early_minutes), 1),
                )
            )

    results.sort(key=lambda r: (-r.early_leave_days, -r.average_early_minutes))
    return results[:limit]


@router.get("/late-detail", response_model=list[LateArrivalEntry])
async def late_detail(
    start_date: date,
    end_date: date,
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    threshold_minutes: int = Query(0, ge=0, le=240, description="Grace period before counting as late"),
    exclude_weekends: bool = True,
    limit: int = Query(1000, ge=1, le=5000),
):
    """Flat, chronological list of every late arrival in the range — one row per
    (staff, day): the day, the clock time of their first IN, the campus shift
    start, and how many minutes late. Sorted by date then time, so it reads as a
    diary of "who came in late, when, and at what time"."""
    _validate_range(start_date, end_date)
    tz = ZoneInfo(settings.attendance_timezone)
    staff = await _scoped_active_staff(db, manager, campus_id, user_id)
    if not staff:
        return []

    logs = await _logs_for_staff(db, [s.id for s in staff], start_date, end_date)
    grouped = _group_by_staff_day(logs, tz)
    campuses = await _campus_shift_map(db)
    national_holidays, holidays_by_campus = await _load_holidays(db, start_date, end_date)
    all_days = _all_days(start_date, end_date)

    rows: list[tuple[datetime, LateArrivalEntry]] = []
    for s in staff:
        campus = campuses.get(s.campus_id) if s.campus_id else None
        if campus is None or campus.shift_start is None:
            continue  # no shift configured for this campus — cannot judge lateness
        for d in _expected_days_for_staff(
            s, all_days, exclude_weekends, national_holidays, holidays_by_campus
        ):
            bucket = grouped.get((s.id, d))
            if bucket is None or bucket.first_in is None:
                continue
            shift_start_local = datetime.combine(d, campus.shift_start, tzinfo=tz)
            minutes_late = (bucket.first_in - shift_start_local).total_seconds() / 60
            if minutes_late > threshold_minutes:
                rows.append(
                    (
                        bucket.first_in,
                        LateArrivalEntry(
                            user_id=s.id,
                            full_name=s.full_name,
                            job_title=s.job_title,
                            branch=s.branch,
                            campus_name=campus.name,
                            date=d,
                            arrival_time=bucket.first_in.strftime("%H:%M"),
                            shift_start=campus.shift_start.strftime("%H:%M"),
                            minutes_late=round(minutes_late),
                        ),
                    )
                )

    rows.sort(key=lambda t: t[0])  # chronological by the actual arrival instant
    return [entry for _, entry in rows[:limit]]


@router.get("/early-leave-detail", response_model=list[EarlyLeaveEntry])
async def early_leave_detail(
    start_date: date,
    end_date: date,
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    threshold_minutes: int = Query(0, ge=0, le=240, description="Grace period before counting as early"),
    exclude_weekends: bool = True,
    limit: int = Query(1000, ge=1, le=5000),
):
    """Flat, chronological list of every early leave in the range — one row per
    (staff, day): the day, the clock time of their last OUT, the campus shift
    end, and how many minutes early. Sorted by date then time."""
    _validate_range(start_date, end_date)
    tz = ZoneInfo(settings.attendance_timezone)
    staff = await _scoped_active_staff(db, manager, campus_id, user_id)
    if not staff:
        return []

    logs = await _logs_for_staff(db, [s.id for s in staff], start_date, end_date)
    grouped = _group_by_staff_day(logs, tz)
    campuses = await _campus_shift_map(db)
    national_holidays, holidays_by_campus = await _load_holidays(db, start_date, end_date)
    all_days = _all_days(start_date, end_date)

    rows: list[tuple[datetime, EarlyLeaveEntry]] = []
    for s in staff:
        campus = campuses.get(s.campus_id) if s.campus_id else None
        if campus is None or campus.shift_end is None:
            continue
        for d in _expected_days_for_staff(
            s, all_days, exclude_weekends, national_holidays, holidays_by_campus
        ):
            bucket = grouped.get((s.id, d))
            if bucket is None or bucket.last_out is None:
                continue
            shift_end_local = datetime.combine(d, campus.shift_end, tzinfo=tz)
            minutes_early = (shift_end_local - bucket.last_out).total_seconds() / 60
            if minutes_early > threshold_minutes:
                rows.append(
                    (
                        bucket.last_out,
                        EarlyLeaveEntry(
                            user_id=s.id,
                            full_name=s.full_name,
                            job_title=s.job_title,
                            branch=s.branch,
                            campus_name=campus.name,
                            date=d,
                            leave_time=bucket.last_out.strftime("%H:%M"),
                            shift_end=campus.shift_end.strftime("%H:%M"),
                            minutes_early=round(minutes_early),
                        ),
                    )
                )

    rows.sort(key=lambda t: t[0])  # chronological by the actual leave instant
    return [entry for _, entry in rows[:limit]]


async def _compute_absences(
    db: AsyncSession,
    manager: User,
    start_date: date,
    end_date: date,
    campus_id: int | None,
    user_id: int | None,
    exclude_weekends: bool,
) -> list[AbsenceDayEntry]:
    _validate_range(start_date, end_date)
    tz = ZoneInfo(settings.attendance_timezone)
    staff = await _scoped_active_staff(db, manager, campus_id, user_id)
    if not staff:
        return []

    all_days = _all_days(start_date, end_date)
    if len(staff) * max(len(all_days), 1) > 8000:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sonuç kümesi çok büyük; tarih aralığını veya kampüs/personel filtresini daraltın.",
        )

    logs = await _logs_for_staff(db, [s.id for s in staff], start_date, end_date)
    present_days = {(log.user_id, log.scan_time.astimezone(tz).date()) for log in logs}

    leave_rows = await db.execute(
        select(LeaveRecord).where(
            LeaveRecord.user_id.in_([s.id for s in staff]),
            LeaveRecord.status == LeaveStatus.active,
            LeaveRecord.start_date <= end_date,
            LeaveRecord.end_date >= start_date,
        )
    )
    leaves_by_staff: dict[int, list[LeaveRecord]] = defaultdict(list)
    for leave in leave_rows.scalars().all():
        leaves_by_staff[leave.user_id].append(leave)

    national_holidays, holidays_by_campus = await _load_holidays(db, start_date, end_date)
    names = await _campus_names_map(db)

    entries: list[AbsenceDayEntry] = []
    for s in staff:
        staff_leaves = leaves_by_staff.get(s.id, [])
        # Only count days this staff member is actually expected to work
        # (their per-person schedule, minus holidays/closures).
        for d in _expected_days_for_staff(
            s, all_days, exclude_weekends, national_holidays, holidays_by_campus
        ):
            if (s.id, d) in present_days:
                continue
            covering = next((lv for lv in staff_leaves if lv.start_date <= d <= lv.end_date), None)
            entries.append(
                AbsenceDayEntry(
                    user_id=s.id,
                    full_name=s.full_name,
                    job_title=s.job_title,
                    branch=s.branch,
                    campus_name=names.get(s.campus_id) if s.campus_id else None,
                    date=d,
                    status=covering.leave_type if covering else "unresolved",
                    leave_record_id=covering.id if covering else None,
                )
            )
    entries.sort(key=lambda e: (e.date, e.full_name))
    return entries


@router.get("/absences", response_model=list[AbsenceDayEntry])
async def absence_detail(
    start_date: date,
    end_date: date,
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    exclude_weekends: bool = True,
):
    """Every (staff, day) with zero scans in the range — labelled with the
    covering leave type, or ``unresolved`` if no status has been entered.
    Days are never silently omitted: an absence without a leave record always
    shows up here as ``unresolved`` so it can't go unnoticed."""
    return await _compute_absences(db, manager, start_date, end_date, campus_id, user_id, exclude_weekends)


@router.get("/absence-summary", response_model=AbsenceSummaryResponse)
async def absence_summary(
    start_date: date,
    end_date: date,
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    exclude_weekends: bool = True,
):
    """Aggregate absence statistics: totals by reason type, and a per-staff
    ranking of most-absent staff (with a reason breakdown each)."""
    entries = await _compute_absences(db, manager, start_date, end_date, campus_id, user_id, exclude_weekends)

    by_reason_days: dict[str, int] = defaultdict(int)
    by_reason_staff: dict[str, set[int]] = defaultdict(set)
    by_staff: dict[int, AbsenceTotalEntry] = {}
    unresolved_count = 0

    for e in entries:
        if e.status == "unresolved":
            unresolved_count += 1
        else:
            by_reason_days[e.status] += 1
            by_reason_staff[e.status].add(e.user_id)

        if e.user_id not in by_staff:
            by_staff[e.user_id] = AbsenceTotalEntry(
                user_id=e.user_id,
                full_name=e.full_name,
                job_title=e.job_title,
                branch=e.branch,
                campus_name=e.campus_name,
                absent_days=0,
                unresolved_days=0,
                by_reason={},
            )
        total = by_staff[e.user_id]
        total.absent_days += 1
        if e.status == "unresolved":
            total.unresolved_days += 1
        else:
            total.by_reason[e.status] = total.by_reason.get(e.status, 0) + 1

    by_reason = [
        AbsenceReasonStat(leave_type=reason, day_count=count, staff_count=len(by_reason_staff[reason]))
        for reason, count in sorted(by_reason_days.items(), key=lambda kv: -kv[1])
    ]
    totals = sorted(by_staff.values(), key=lambda t: -t.absent_days)

    return AbsenceSummaryResponse(
        start_date=start_date,
        end_date=end_date,
        by_reason=by_reason,
        totals_by_staff=totals,
        unresolved_count=unresolved_count,
    )


@router.get("/risk", response_model=RiskReportResponse)
async def risk_report(
    start_date: date,
    end_date: date,
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    threshold_minutes: int = Query(0, ge=0, le=240, description="Grace before late/early counts"),
    exclude_weekends: bool = True,
    late_threshold: int = Query(3, ge=1, le=100, description="Flag at this many late days"),
    early_leave_threshold: int = Query(3, ge=1, le=100, description="Flag at this many early-leave days"),
    unresolved_threshold: int = Query(2, ge=1, le=100, description="Flag at this many unresolved-absence days"),
):
    """Early-warning panel: surfaces staff whose late arrivals, early leaves, or
    unresolved (durum girilmemiş) absences in the range cross the configured
    thresholds. Pure threshold layer over the existing late/early-leave/absence
    computations — no new data source, so it always agrees with the detail
    reports."""
    late = await late_ranking(
        start_date, end_date, manager, db, campus_id, user_id, threshold_minutes, exclude_weekends, 700
    )
    early = await early_leave_ranking(
        start_date, end_date, manager, db, campus_id, user_id, threshold_minutes, exclude_weekends, 700
    )
    absences = await _compute_absences(db, manager, start_date, end_date, campus_id, user_id, exclude_weekends)

    # Aggregate per staff, carrying display fields from whichever source has them.
    agg: dict[int, dict] = {}

    def _slot(uid, full_name, job_title, branch, campus_name) -> dict:
        s = agg.get(uid)
        if s is None:
            s = {
                "user_id": uid,
                "full_name": full_name,
                "job_title": job_title,
                "branch": branch,
                "campus_name": campus_name,
                "late_days": 0,
                "early_leave_days": 0,
                "unresolved_days": 0,
            }
            agg[uid] = s
        return s

    for r in late:
        _slot(r.user_id, r.full_name, r.job_title, r.branch, r.campus_name)["late_days"] = r.late_days
    for r in early:
        _slot(r.user_id, r.full_name, r.job_title, r.branch, r.campus_name)["early_leave_days"] = r.early_leave_days
    for e in absences:
        if e.status == "unresolved":
            _slot(e.user_id, e.full_name, e.job_title, e.branch, e.campus_name)["unresolved_days"] += 1

    entries: list[RiskStaffEntry] = []
    for s in agg.values():
        flags: list[str] = []
        if s["late_days"] >= late_threshold:
            flags.append(f"Bu dönem {s['late_days']} kez geç geldi")
        if s["early_leave_days"] >= early_leave_threshold:
            flags.append(f"Bu dönem {s['early_leave_days']} kez erken çıktı")
        if s["unresolved_days"] >= unresolved_threshold:
            flags.append(f"{s['unresolved_days']} gün durumu girilmemiş devamsızlık")
        if not flags:
            continue  # early-warning list: only flagged staff appear

        score = s["unresolved_days"] * 3 + s["late_days"] + s["early_leave_days"]
        # "high" when the problem is serious: any unresolved absence over the
        # bar, a doubled late/early pattern, or more than one kind of flag.
        is_high = (
            s["unresolved_days"] >= unresolved_threshold
            or s["late_days"] >= 2 * late_threshold
            or s["early_leave_days"] >= 2 * early_leave_threshold
            or len(flags) >= 2
        )
        entries.append(
            RiskStaffEntry(
                user_id=s["user_id"],
                full_name=s["full_name"],
                job_title=s["job_title"],
                branch=s["branch"],
                campus_name=s["campus_name"],
                late_days=s["late_days"],
                early_leave_days=s["early_leave_days"],
                unresolved_days=s["unresolved_days"],
                score=score,
                level="high" if is_high else "medium",
                flags=flags,
            )
        )

    entries.sort(key=lambda e: (0 if e.level == "high" else 1, -e.score, e.full_name))
    return RiskReportResponse(
        start_date=start_date,
        end_date=end_date,
        high_count=sum(1 for e in entries if e.level == "high"),
        medium_count=sum(1 for e in entries if e.level == "medium"),
        entries=entries,
        late_threshold=late_threshold,
        early_leave_threshold=early_leave_threshold,
        unresolved_threshold=unresolved_threshold,
    )


@router.get("/unresolved-reminder", response_model=UnresolvedReminderResponse)
async def unresolved_reminder(
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    days: int = Query(14, ge=1, le=90, description="Trailing window length (ending yesterday)"),
    exclude_weekends: bool = True,
):
    """Absence days in the trailing window that still have no status entered —
    surfaced so a manager is reminded to resolve them (mark a leave or correct
    the record) instead of letting an unexplained gap slip by unnoticed.

    The window ends *yesterday* (local): today is still in progress, so its
    missing scans are not yet a real absence."""
    tz = ZoneInfo(settings.attendance_timezone)
    today_local = datetime.now(timezone.utc).astimezone(tz).date()
    end_date = today_local - timedelta(days=1)
    start_date = end_date - timedelta(days=days - 1)
    if end_date < start_date:  # days == 0 guard (already bounded ≥1)
        start_date = end_date

    entries = await _compute_absences(
        db, manager, start_date, end_date, campus_id, None, exclude_weekends
    )
    unresolved = [e for e in entries if e.status == "unresolved"]
    return UnresolvedReminderResponse(
        start_date=start_date,
        end_date=end_date,
        unresolved_count=len(unresolved),
        entries=unresolved,
    )


@router.get("/daily-trend", response_model=DailyTrendResponse)
async def daily_trend(
    start_date: date,
    end_date: date,
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    exclude_weekends: bool = True,
):
    """Per-day attendance aggregates over the range: how many staff were
    expected, present, on leave, or absent-without-status (unresolved). Drives
    the dashboard/report trend chart — narrowed to one staff member when
    ``user_id`` is given, so it doubles as their personal daily/weekly/monthly
    trend."""
    _validate_range(start_date, end_date)
    tz = ZoneInfo(settings.attendance_timezone)
    staff = await _scoped_active_staff(db, manager, campus_id, user_id)
    all_days = _all_days(start_date, end_date)

    # Per-day accumulators, seeded so every day in the range is present even if
    # it is a weekend/holiday (expected == 0 there).
    buckets: dict[date, DailyTrendEntry] = {
        d: DailyTrendEntry(date=d, expected=0, present=0, on_leave=0, unresolved=0)
        for d in all_days
    }

    if staff:
        if len(staff) * max(len(all_days), 1) > 20000:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Sonuç kümesi çok büyük; tarih aralığını veya kampüs filtresini daraltın.",
            )
        logs = await _logs_for_staff(db, [s.id for s in staff], start_date, end_date)
        present_days = {(log.user_id, log.scan_time.astimezone(tz).date()) for log in logs}

        leave_rows = await db.execute(
            select(LeaveRecord).where(
                LeaveRecord.user_id.in_([s.id for s in staff]),
                LeaveRecord.status == LeaveStatus.active,
                LeaveRecord.start_date <= end_date,
                LeaveRecord.end_date >= start_date,
            )
        )
        leaves_by_staff: dict[int, list[LeaveRecord]] = defaultdict(list)
        for leave in leave_rows.scalars().all():
            leaves_by_staff[leave.user_id].append(leave)

        national_holidays, holidays_by_campus = await _load_holidays(db, start_date, end_date)

        for s in staff:
            staff_leaves = leaves_by_staff.get(s.id, [])
            for d in _expected_days_for_staff(
                s, all_days, exclude_weekends, national_holidays, holidays_by_campus
            ):
                entry = buckets[d]
                entry.expected += 1
                if (s.id, d) in present_days:
                    entry.present += 1
                elif any(lv.start_date <= d <= lv.end_date for lv in staff_leaves):
                    entry.on_leave += 1
                else:
                    entry.unresolved += 1

    entries = [buckets[d] for d in all_days]
    return DailyTrendResponse(
        start_date=start_date,
        end_date=end_date,
        entries=entries,
        total_expected=sum(e.expected for e in entries),
        total_present=sum(e.present for e in entries),
        total_on_leave=sum(e.on_leave for e in entries),
        total_unresolved=sum(e.unresolved for e in entries),
    )


@router.get("/forgot-checkout", response_model=ForgotCheckoutResponse)
async def forgot_checkout(
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    threshold_minutes: int = Query(30, ge=0, le=240, description="Grace after shift end"),
):
    """Staff still 'inside' right now (their last log today is an IN) whose
    campus shift has already ended by more than ``threshold_minutes`` — they
    most likely forgot to scan out and will be auto-closed at 23:59. Surfaced
    so a manager can nudge them before that happens."""
    now_utc = datetime.now(timezone.utc)
    tz = ZoneInfo(settings.attendance_timezone)
    start_utc, end_utc = (
        datetime.combine(now_utc.astimezone(tz).date(), time.min, tzinfo=tz).astimezone(timezone.utc),
        datetime.combine(now_utc.astimezone(tz).date(), time.max, tzinfo=tz).astimezone(timezone.utc),
    )
    scope = scope_campus_id(manager, campus_id)

    latest_subq = (
        select(
            AttendanceLog.user_id.label("uid"),
            func.max(AttendanceLog.scan_time).label("max_time"),
        )
        .where(AttendanceLog.scan_time >= start_utc, AttendanceLog.scan_time <= end_utc)
        .group_by(AttendanceLog.user_id)
        .subquery()
    )
    stmt = (
        select(AttendanceLog, User, Campus)
        .join(
            latest_subq,
            (AttendanceLog.user_id == latest_subq.c.uid)
            & (AttendanceLog.scan_time == latest_subq.c.max_time),
        )
        .join(User, User.id == AttendanceLog.user_id)
        .join(Campus, Campus.id == User.campus_id, isouter=True)
        .where(AttendanceLog.type == AttendanceType.IN)
    )
    if scope is not None:
        stmt = stmt.where(User.campus_id == scope)

    rows = await db.execute(stmt)
    entries: list[ForgotCheckoutEntry] = []
    for log, user, campus in rows.all():
        if campus is None or campus.shift_end is None:
            continue  # no shift configured → cannot judge "overdue"
        shift_end_local = datetime.combine(
            now_utc.astimezone(tz).date(), campus.shift_end, tzinfo=tz
        )
        minutes_overdue = (now_utc - shift_end_local.astimezone(timezone.utc)).total_seconds() / 60
        if minutes_overdue <= threshold_minutes:
            continue
        entries.append(
            ForgotCheckoutEntry(
                user_id=user.id,
                full_name=user.full_name,
                campus_name=campus.name,
                since=log.scan_time,
                minutes_overdue=int(minutes_overdue),
            )
        )
    entries.sort(key=lambda e: -e.minutes_overdue)
    return ForgotCheckoutResponse(as_of=now_utc, entries=entries)


@router.get("/location-alerts", response_model=LocationAlertsResponse)
async def location_alerts(
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    limit: int = Query(200, ge=1, le=1000),
):
    """Far-from-campus QR scan attempts (geofence violations): a staff member
    tried to scan in/out while their phone was outside the campus radius. The
    attendance was rejected; each attempt is listed here for the manager."""
    tz = ZoneInfo(settings.attendance_timezone)
    scope = scope_campus_id(manager, campus_id)

    stmt = (
        select(LocationViolation, User, Campus)
        .join(User, User.id == LocationViolation.user_id)
        .join(Campus, Campus.id == LocationViolation.campus_id, isouter=True)
    )
    if scope is not None:
        stmt = stmt.where(LocationViolation.campus_id == scope)
    if start_date is not None:
        start_utc = datetime.combine(start_date, time.min, tzinfo=tz).astimezone(timezone.utc)
        stmt = stmt.where(LocationViolation.created_at >= start_utc)
    if end_date is not None:
        end_utc = datetime.combine(end_date, time.max, tzinfo=tz).astimezone(timezone.utc)
        stmt = stmt.where(LocationViolation.created_at <= end_utc)
    stmt = stmt.order_by(LocationViolation.created_at.desc()).limit(limit)

    rows = await db.execute(stmt)
    entries: list[LocationAlertEntry] = []
    for v, user, campus in rows.all():
        entries.append(
            LocationAlertEntry(
                id=v.id,
                user_id=user.id,
                full_name=user.full_name,
                job_title=user.job_title,
                branch=user.branch,
                campus_name=campus.name if campus else None,
                distance_m=round(v.distance_m),
                accuracy_m=round(v.accuracy_m) if v.accuracy_m is not None else None,
                latitude=v.latitude,
                longitude=v.longitude,
                maps_url=f"https://www.google.com/maps?q={v.latitude},{v.longitude}",
                created_at=v.created_at,
            )
        )
    return LocationAlertsResponse(count=len(entries), entries=entries)


@router.get("/monthly-hours", response_model=MonthlyHoursResponse)
async def monthly_hours(
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    exclude_weekends: bool = True,
):
    """Per-staff monthly attendance totals for payroll (puantaj): worked hours,
    days present, cumulative late minutes, and absent/leave day counts over the
    days each person was scheduled to work.

    Worked hours are the sum of (last OUT − first IN) per day. Note that on a
    day where someone forgot to scan out, the 23:59 auto-close OUT stands in as
    their last OUT, so that day's hours run long — ``worked_days`` vs
    ``present_days`` surfaces such gaps.
    """
    start_date = date(year, month, 1)
    end_date = date(year, month, calendar.monthrange(year, month)[1])
    tz = ZoneInfo(settings.attendance_timezone)

    staff = await _scoped_active_staff(db, manager, campus_id, user_id)
    if not staff:
        return MonthlyHoursResponse(
            year=year, month=month, start_date=start_date, end_date=end_date, entries=[]
        )

    logs = await _logs_for_staff(db, [s.id for s in staff], start_date, end_date)
    grouped = _group_by_staff_day(logs, tz)
    grouped_by_user: dict[int, list[tuple[date, _DayLogs]]] = defaultdict(list)
    for (uid, d), bucket in grouped.items():
        grouped_by_user[uid].append((d, bucket))

    campuses = await _campus_shift_map(db)
    national_holidays, holidays_by_campus = await _load_holidays(db, start_date, end_date)
    all_days = _all_days(start_date, end_date)
    names = await _campus_names_map(db)

    leave_rows = await db.execute(
        select(LeaveRecord).where(
            LeaveRecord.user_id.in_([s.id for s in staff]),
            LeaveRecord.status == LeaveStatus.active,
            LeaveRecord.start_date <= end_date,
            LeaveRecord.end_date >= start_date,
        )
    )
    leaves_by_staff: dict[int, list[LeaveRecord]] = defaultdict(list)
    for leave in leave_rows.scalars().all():
        leaves_by_staff[leave.user_id].append(leave)

    entries: list[MonthlyHoursEntry] = []
    for s in staff:
        campus = campuses.get(s.campus_id) if s.campus_id else None
        expected = _expected_days_for_staff(
            s, all_days, exclude_weekends, national_holidays, holidays_by_campus
        )
        staff_leaves = leaves_by_staff.get(s.id, [])

        # Worked hours / present / complete-day counts use every day with scans
        # (someone may have worked an unscheduled day too).
        total_seconds = 0.0
        worked_days = 0
        present_days = 0
        for _d, bucket in grouped_by_user.get(s.id, []):
            if bucket.any_log:
                present_days += 1
            if bucket.first_in and bucket.last_out and bucket.last_out > bucket.first_in:
                total_seconds += (bucket.last_out - bucket.first_in).total_seconds()
                worked_days += 1

        # Lateness + absent/leave classification only over scheduled days.
        total_late = 0.0
        absent_days = 0
        leave_days = 0
        for d in expected:
            bucket = grouped.get((s.id, d))
            if bucket and bucket.any_log:
                if campus and campus.shift_start and bucket.first_in:
                    shift_start_local = datetime.combine(d, campus.shift_start, tzinfo=tz)
                    minutes_late = (bucket.first_in - shift_start_local).total_seconds() / 60
                    if minutes_late > 0:
                        total_late += minutes_late
                continue
            covering = next((lv for lv in staff_leaves if lv.start_date <= d <= lv.end_date), None)
            if covering:
                leave_days += 1
            else:
                absent_days += 1

        entries.append(
            MonthlyHoursEntry(
                user_id=s.id,
                full_name=s.full_name,
                job_title=s.job_title,
                branch=s.branch,
                campus_name=names.get(s.campus_id) if s.campus_id else None,
                expected_days=len(expected),
                present_days=present_days,
                worked_days=worked_days,
                total_hours=round(total_seconds / 3600, 1),
                total_late_minutes=round(total_late),
                absent_days=absent_days,
                leave_days=leave_days,
            )
        )

    entries.sort(key=lambda e: (e.campus_name or "", e.full_name))
    return MonthlyHoursResponse(
        year=year, month=month, start_date=start_date, end_date=end_date, entries=entries
    )


@router.get("/monthly-hours.xlsx", response_class=StreamingResponse)
async def export_monthly_hours_xlsx(
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    exclude_weekends: bool = True,
):
    """The monthly-hours (puantaj) report as a single-sheet workbook."""
    report = await monthly_hours(year, month, manager, db, campus_id, user_id, exclude_weekends)

    wb = Workbook()
    ws = wb.active
    ws.title = "Aylık Mesai"
    ws.append(
        [
            "Personel", "Görev", "Branş", "Kampüs",
            "Planlı Gün", "Geldiği Gün", "Tam Gün (Giriş+Çıkış)",
            "Toplam Saat", "Toplam Geç (dk)", "Devamsız Gün", "İzinli Gün",
        ]
    )
    for e in report.entries:
        ws.append(
            [
                e.full_name, e.job_title or "", e.branch or "", e.campus_name or "",
                e.expected_days, e.present_days, e.worked_days,
                e.total_hours, e.total_late_minutes, e.absent_days, e.leave_days,
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"aylik_mesai_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export.xlsx", response_class=StreamingResponse)
async def export_reports_xlsx(
    start_date: date,
    end_date: date,
    manager: User = Depends(get_current_manager),
    db: AsyncSession = Depends(get_db),
    campus_id: int | None = Query(None, description="hq only: filter to one campus"),
    user_id: int | None = None,
    threshold_minutes: int = Query(0, ge=0, le=240),
    exclude_weekends: bool = True,
):
    """One workbook with a sheet each for late/early-leave rankings and the
    absence detail + summary, for the given date range and scope — narrowed to
    one staff member when ``user_id`` is given."""
    late = await late_ranking(
        start_date, end_date, manager, db, campus_id, user_id, threshold_minutes, exclude_weekends, 700
    )
    early = await early_leave_ranking(
        start_date, end_date, manager, db, campus_id, user_id, threshold_minutes, exclude_weekends, 700
    )
    late_list = await late_detail(
        start_date, end_date, manager, db, campus_id, user_id, threshold_minutes, exclude_weekends, 5000
    )
    early_list = await early_leave_detail(
        start_date, end_date, manager, db, campus_id, user_id, threshold_minutes, exclude_weekends, 5000
    )
    absences = await _compute_absences(db, manager, start_date, end_date, campus_id, user_id, exclude_weekends)
    summary = await absence_summary(start_date, end_date, manager, db, campus_id, user_id, exclude_weekends)

    wb = Workbook()

    ws = wb.active
    ws.title = "Geç Kalmalar"
    ws.append(["Personel", "Görev", "Branş", "Kampüs", "Geç Kaldığı Gün Sayısı", "Ortalama Geç Kalma (dk)"])
    for r in late:
        ws.append([r.full_name, r.job_title or "", r.branch or "", r.campus_name or "", r.late_days, r.average_late_minutes])

    ws2 = wb.create_sheet("Erken Çıkışlar")
    ws2.append(["Personel", "Görev", "Branş", "Kampüs", "Erken Çıktığı Gün Sayısı", "Ortalama Erken Çıkma (dk)"])
    for r in early:
        ws2.append([r.full_name, r.job_title or "", r.branch or "", r.campus_name or "", r.early_leave_days, r.average_early_minutes])

    ws_ld = wb.create_sheet("Geç Giriş Listesi")
    ws_ld.append(["Tarih", "Saat", "Personel", "Görev", "Branş", "Kampüs", "Mesai Başlangıcı", "Gecikme (dk)"])
    for e in late_list:
        ws_ld.append(
            [e.date.isoformat(), e.arrival_time, e.full_name, e.job_title or "", e.branch or "", e.campus_name or "", e.shift_start, e.minutes_late]
        )

    ws_ed = wb.create_sheet("Erken Çıkış Listesi")
    ws_ed.append(["Tarih", "Saat", "Personel", "Görev", "Branş", "Kampüs", "Mesai Bitişi", "Erken (dk)"])
    for e in early_list:
        ws_ed.append(
            [e.date.isoformat(), e.leave_time, e.full_name, e.job_title or "", e.branch or "", e.campus_name or "", e.shift_end, e.minutes_early]
        )

    ws3 = wb.create_sheet("Devamsızlık Detay")
    ws3.append(["Personel", "Görev", "Branş", "Kampüs", "Tarih", "Durum"])
    for e in absences:
        ws3.append([e.full_name, e.job_title or "", e.branch or "", e.campus_name or "", e.date.isoformat(), e.status])

    ws4 = wb.create_sheet("Devamsızlık Özeti")
    ws4.append(["İzin/Durum Türü", "Toplam Gün", "Personel Sayısı"])
    for r in summary.by_reason:
        ws4.append([r.leave_type, r.day_count, r.staff_count])
    ws4.append([])
    ws4.append(["Durum Girilmemiş Gün Sayısı", summary.unresolved_count])
    ws4.append([])
    ws4.append(["Personel", "Görev", "Branş", "Kampüs", "Toplam Devamsız Gün", "Durum Girilmemiş Gün"])
    for t in summary.totals_by_staff:
        ws4.append([t.full_name, t.job_title or "", t.branch or "", t.campus_name or "", t.absent_days, t.unresolved_days])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"raporlar_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
